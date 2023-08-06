import openpyxl
import re
import os
from datetime import date as dt, timedelta


class Excel:

    def __init__(self, directory, name):
        self.filename = name
        self.filePath = os.path.join(directory, name)
        self.wb = openpyxl.load_workbook(self.filePath)
        self.headers = {}

    def getFileName(self):
        return self.filename

    def getPath(self):
        return self.filePath

    def save(self, path):
        self.wb.save(path)

    def getHeaders(self):
        ws = self.wb.active
        self.headers["A1"] = "Attività istituzionale"
        for row in ws.iter_rows(max_col=1):
            for cell in row:
                m = re.match("Progetto \d+", str(cell.value))
                if m is not None:
                    self.headers[cell.coordinate] = m.group()
        return self.headers

    def __getMonthDays(self, year, m, ws):
        if m == "01":
            for months in ws.iter_rows(min_col=2, max_col=2):
                for cell in months:
                    if cell.value == "GENNAIO":
                        start = (cell.row + 1, cell.column)
                        end = cell.column + 30
                        return start, end
        if m == "02":
            for months in ws.iter_rows(min_col=2, max_col=2):
                for cell in months:
                    if cell.value == "FEBBRAIO":
                        start = (cell.row + 1, cell.column)
                        if year % 4 == 0 and year % 100 != 0 or year % 400 == 0:
                            end = cell.column + 28
                        else:
                            end = cell.column + 27
                        return start, end
        if m == "03":
            for months in ws.iter_rows(min_col=2, max_col=2):
                for cell in months:
                    if cell.value == "MARZO":
                        start = (cell.row + 1, cell.column)
                        end = cell.column + 30
                        return start, end
        if m == "04":
            for months in ws.iter_rows(min_col=2, max_col=2):
                for cell in months:
                    if cell.value == "APRILE":
                        start = (cell.row + 1, cell.column)
                        end = cell.column + 29
                        return start, end
        if m == "05":
            for months in ws.iter_rows(min_col=2, max_col=2):
                for cell in months:
                    if cell.value == "MAGGIO":
                        start = (cell.row + 1, cell.column)
                        end = cell.column + 30
                        return start, end
        if m == "06":
            for months in ws.iter_rows(min_col=2, max_col=2):
                for cell in months:
                    if cell.value == "GIUGNO":
                        start = (cell.row + 1, cell.column)
                        end = cell.column + 29
                        return start, end
        if m == "07":
            for months in ws.iter_rows(min_col=2, max_col=2):
                for cell in months:
                    if cell.value == "LUGLIO":
                        start = (cell.row + 1, cell.column)
                        end = cell.column + 30
                        return start, end
        if m == "08":
            for months in ws.iter_rows(min_col=2, max_col=2):
                for cell in months:
                    if cell.value == "AGOSTO":
                        start = (cell.row + 1, cell.column)
                        end = cell.column + 30
                        return start, end
        if m == "09":
            for months in ws.iter_rows(min_col=2, max_col=2):
                for cell in months:
                    if cell.value == "SETTEMBRE":
                        start = (cell.row + 1, cell.column)
                        end = cell.column + 29
                        return start, end
        if m == "10":
            for months in ws.iter_rows(min_col=2, max_col=2):
                for cell in months:
                    if cell.value == "OTTOBRE":
                        start = (cell.row + 1, cell.column)
                        end = cell.column + 30
                        return start, end
        if m == "11":
            for months in ws.iter_rows(min_col=2, max_col=2):
                for cell in months:
                    if cell.value == "NOVEMBRE":
                        start = (cell.row + 1, cell.column)
                        end = cell.column + 29
                        return start, end
        if m == "12":
            for months in ws.iter_rows(min_col=2, max_col=2):
                for cell in months:
                    if cell.value == "DICEMBRE":
                        start = (cell.row + 1, cell.column)
                        end = cell.column + 30
                        return start, end

    def modify(self, activity, data, nore, multiplier):
        numOre = nore
        day, month, year = data.split("/")
        date = dt(int(year), int(month), int(day))
        ws = self.wb[year]
        start, end = self.__getMonthDays(int(year), month, ws)
        for days in ws.iter_rows(min_row=start[0], max_row=start[0], min_col=start[1], max_col=end):
            for cell2 in days:
                if cell2.value == int(day):
                    for activities in ws.iter_rows(min_row=cell2.row + 1, max_col=1):
                        for cell3 in activities:
                            value = None
                            for keys in self.headers:
                                if self.headers[keys] == activity:
                                    value = "=" + keys
                            if cell3.value == value or cell3.value == activity:
                                point = cell2.column_letter + str(cell3.row)
                                if ws[point].value is not None:
                                    ws[point] = ws[point].value + round(numOre * multiplier)
                                else:
                                    ws[point] = round(numOre * multiplier)
                                if ws[point].value > 7:
                                    if not date.month == date.day == 1 and date.year == int(year):
                                        plus = ws[point].value - 7
                                        ws[point] = 7
                                        new_row = cell3.row
                                        new_col = cell2.column - 1
                                        new_date = date - timedelta(days=1)
                                        while plus > 0 and (new_date.year == date.year):
                                            if new_date.month == date.month - 1:
                                                null, new_col = self.__getMonthDays(int(year), new_date.strftime("%m"),
                                                                                    ws)
                                                new_row = null[0] + (cell3.row - start[0])
                                                date = new_date
                                            if new_date.weekday() != 6:
                                                if ws.cell(new_row, new_col).value is not None:
                                                    ws.cell(new_row, new_col).value += 1
                                                else:
                                                    ws.cell(new_row, new_col).value = 1
                                                plus -= 1
                                                if ws.cell(new_row, new_col).value == 7:
                                                    if not new_date.month == new_date.day == 1:
                                                        new_col -= 1
                                                        new_date = new_date - timedelta(days=1)
                                            else:
                                                new_col -= 1
                                                new_date = new_date - timedelta(days=1)
                                        if plus > 0:
                                            if ws.cell(new_row, new_col).value is not None:
                                                ws.cell(new_row, new_col).value += plus
                                            else:
                                                ws.cell(new_row, new_col).value = plus
                                return
