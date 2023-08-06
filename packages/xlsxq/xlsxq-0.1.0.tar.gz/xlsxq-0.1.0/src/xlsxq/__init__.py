from enum import Enum
from openpyxl import load_workbook
from pathlib import Path
import argparse
import json
import sys


__version__ = '0.1.0'
ARG_JSON = 'json'
ARG_TSV = 'tsv'


class SheetNotFoundError(RuntimeError):
    pass


class InternalError(RuntimeError):
    pass


class QueryNameEnum(Enum):
    SHEET_LIST = 1
    RANGE_SHOW = 2


class OutputNameEnum(Enum):
    JSON = ARG_JSON
    TSV = ARG_TSV


class Query(object):
    def execute(self):
        raise NotImplementedError()


class Sheet(object):
    def __init__(self, name: str):
        self._name = name

    @property
    def name(self):
        return self._name


class Dumper(object):
    def dump(self, obj, f=sys.stdout):
        raise NotImplementedError()


class JsonDumper(Dumper):
    def __init__(self):
        self._indent = 2

    def dump(self, obj, f=sys.stdout):
        class JsonEncoder(json.JSONEncoder):
            def default(self, obj):
                if isinstance(obj, Sheet):
                    return {'name': obj.name}
                return json.JSONEncoder.default(self, obj)
        f.write(json.dumps(obj, indent=self._indent, cls=JsonEncoder))


class TsvDumper(Dumper):
    def __init__(self):
        self._indent = 2

    def dump(self, obj, f=sys.stdout):
        if not isinstance(obj, list):
            raise InternalError()

        class ValueConverter(object):
            def to_str(self, value):
                if value is None:
                    return ''
                if isinstance(value, str):
                    return value
                if isinstance(value, int):
                    return str(value)
                raise NotImplementedError(
                        'Type %s is not implemented.' % (type(value)))
        conv = ValueConverter()

        text = ''
        for row in obj:
            if not isinstance(row, list):
                raise InternalError()
            text += "\t".join(map(lambda x: conv.to_str(x), row)) + "\n"
        f.write(text)


class DumperFactory(object):
    def __init__(self):
        self._name_cls = {
            ARG_JSON: JsonDumper,
            ARG_TSV: TsvDumper
        }

    def create(self, name):
        cls = self._name_cls[name]
        return cls()


class SheetListQuery(Query):
    def __init__(self, infile: str, output: str):
        self._infile = Path(infile)
        if not self._infile.exists():
            raise FileNotFoundError(f'File {self._infile} was not found.')
        self._output = output

    def execute(self):
        book = load_workbook(str(self._infile))

        sheet_list: list[Sheet] = list()
        for name in book.sheetnames:
            sheet_list.append(Sheet(name))

        dumper = DumperFactory().create(self._output)
        dumper.dump(sheet_list)


class RangeShowQuery(Query):
    def __init__(self, infile: str, sheet: str, range_: str, output: str):
        self._infile = Path(infile)
        if not self._infile.exists():
            raise FileNotFoundError(f'File {self._infile} was not found.')
        self._sheet = sheet
        self._range = range_
        self._output = output

    def execute(self):
        book = load_workbook(str(self._infile))

        if self._sheet not in book:
            raise SheetNotFoundError(f'Sheet {self._sheet} was not found.')
        sheet = book[self._sheet]
        cell_range = sheet[self._range]

        table = list()
        for range_row in cell_range:
            row = list()
            for cell in range_row:
                row.append(cell.value)
            table.append(row)

        dumper = DumperFactory().create(self._output)
        dumper.dump(table)


class QueryFactory(object):
    def __init__(self):
        self._name_cls = {
            QueryNameEnum.SHEET_LIST: SheetListQuery,
            QueryNameEnum.RANGE_SHOW: RangeShowQuery
        }

    def create(self, name: str, kwargs: dict):
        cls = self._name_cls[name]
        return cls(**kwargs)


def parse_arguments(args=None):
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers()

    parser_sheet = subparsers.add_parser('sheet')
    subparsers_sheet = parser_sheet.add_subparsers()

    parser_sheet_list = subparsers_sheet.add_parser('list')
    parser_sheet_list.set_defaults(query_name=QueryNameEnum.SHEET_LIST)
    parser_sheet_list.add_argument('--infile', required=True)
    parser_sheet_list.add_argument(
            '--output', choices=[ARG_JSON], required=True)

    parser_range = subparsers.add_parser('range')
    subparsers_range = parser_range.add_subparsers()

    parser_range_show = subparsers_range.add_parser('show')
    parser_range_show.set_defaults(query_name=QueryNameEnum.RANGE_SHOW)
    parser_range_show.add_argument('--infile', required=True)
    parser_range_show.add_argument('--sheet', required=True)
    parser_range_show.add_argument('--range', required=True)
    parser_range_show.add_argument(
            '--output', choices=[ARG_JSON, ARG_TSV], required=True)

    return parser.parse_args(args)


def main():
    ns = parse_arguments()

    query_factory = QueryFactory()
    if ns.query_name == QueryNameEnum.SHEET_LIST:
        kwargs = {
            'infile': ns.infile,
            'output': ns.output
        }
    elif ns.query_name == QueryNameEnum.RANGE_SHOW:
        kwargs = {
            'infile': ns.infile,
            'sheet': ns.sheet,
            'range_': ns.range,
            'output': ns.output
        }
    else:
        raise ValueError()
    query = query_factory.create(ns.query_name, kwargs)
    query.execute()


if __name__ == '__main__':
    main()
